from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.auth import get_current_user
from app.models.models import User, Student, Subject, Attendance, Certification
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

router = APIRouter(prefix="/reports", tags=["Reports"])

def calc_cgpa(subjects):
    passed = [s for s in subjects if s.grade != "F"]
    if not passed: return 0.0
    tc = sum(s.credits for s in passed)
    wp = sum(s.grade_points * s.credits for s in passed)
    return round(wp / tc, 2) if tc > 0 else 0.0

@router.get("/pdf")
def download_pdf_report(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    student = db.query(Student).filter(Student.user_id == current_user.id).first()
    if not student: raise HTTPException(status_code=404, detail="Profile not found")
    subjects = db.query(Subject).filter(Subject.student_id == student.id).order_by(Subject.semester).all()
    attendance = db.query(Attendance).filter(Attendance.student_id == student.id).all()
    certs = db.query(Certification).filter(Certification.student_id == student.id).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []

    title_s = ParagraphStyle('t', fontSize=20, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=6)
    sub_s = ParagraphStyle('s', fontSize=11, alignment=TA_CENTER, spaceAfter=4, textColor=colors.grey)
    info_s = ParagraphStyle('i', fontSize=11, spaceAfter=4)
    h2_s = ParagraphStyle('h2', fontSize=13, fontName='Helvetica-Bold', spaceAfter=6)
    h3_s = ParagraphStyle('h3', fontSize=11, fontName='Helvetica-Bold', spaceAfter=4, textColor=colors.HexColor('#534AB7'))

    elements.append(Paragraph("Student Performance Report", title_s))
    elements.append(Paragraph("Analytics System", sub_s))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"<b>Name:</b> {student.full_name}", info_s))
    elements.append(Paragraph(f"<b>Roll:</b> {student.roll_number}", info_s))
    elements.append(Paragraph(f"<b>Degree:</b> {student.degree} — {student.branch}", info_s))
    elements.append(Paragraph(f"<b>Semester:</b> {student.current_semester}", info_s))
    elements.append(Paragraph(f"<b>CGPA:</b> {calc_cgpa(subjects)}", info_s))
    elements.append(Spacer(1, 0.2 * inch))

    if subjects:
        elements.append(Paragraph("<b>Academic Performance</b>", h2_s))
        for sem in sorted(set(s.semester for s in subjects)):
            sem_subs = [s for s in subjects if s.semester == sem]
            elements.append(Paragraph(f"Semester {sem}", h3_s))
            td = [["Subject", "Code", "Credits", "Internal", "External", "Total", "Grade"]]
            for s in sem_subs:
                td.append([s.name[:30], s.code or "-", str(s.credits), str(s.internal_marks), str(s.external_marks), str(s.total_marks), s.grade or "-"])
            passed = [s for s in sem_subs if s.grade != "F"]
            if passed:
                tc = sum(s.credits for s in passed)
                wp = sum(s.grade_points * s.credits for s in passed)
                td.append(["", "", "", "", "", f"SGPA: {round(wp/tc,2)}", ""])
            t = Table(td, colWidths=[2.2*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.6*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#534AB7')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTSIZE', (0,1), (-1,-1), 9),
                ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F5F5FF')]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 0.15 * inch))

    if attendance:
        elements.append(Paragraph("<b>Attendance Summary</b>", h2_s))
        ad = [["Subject", "Total", "Attended", "%", "Status"]]
        for a in attendance:
            subj = db.query(Subject).filter(Subject.id == a.subject_id).first()
            name = subj.name if subj else f"#{a.subject_id}"
            ad.append([name[:25], str(a.total_classes), str(a.attended_classes), f"{a.percentage}%", "Shortage" if a.is_shortage else "OK"])
        at = Table(ad, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        at.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1D9E75')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ]))
        elements.append(at)

    if certs:
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph("<b>Certifications</b>", h2_s))
        cd = [["Certificate", "Platform", "Skills", "Date"]]
        for c in certs:
            cd.append([c.name[:30], c.platform or "-", (c.skills_gained or "-")[:40], str(c.completed_date.strftime("%b %Y")) if c.completed_date else "-"])
        ct = Table(cd, colWidths=[2*inch, 1.2*inch, 2*inch, 1*inch])
        ct.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#BA7517')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ]))
        elements.append(ct)

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={student.full_name}_report.pdf"})

@router.get("/excel")
def download_excel_report(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    student = db.query(Student).filter(Student.user_id == current_user.id).first()
    if not student: raise HTTPException(status_code=404, detail="Profile not found")
    subjects = db.query(Subject).filter(Subject.student_id == student.id).order_by(Subject.semester).all()
    attendance = db.query(Attendance).filter(Attendance.student_id == student.id).all()
    certs = db.query(Certification).filter(Certification.student_id == student.id).all()

    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Grades
    ws1 = wb.active
    ws1.title = "Grades"
    fill1 = PatternFill(start_color="534AB7", end_color="534AB7", fill_type="solid")
    for col, h in enumerate(["Semester","Subject","Code","Credits","Internal","External","Total","Grade","GP"], 1):
        c = ws1.cell(row=1, column=col, value=h); c.fill = fill1; c.font = hf; c.alignment = center
    for row, s in enumerate(subjects, 2):
        for col, v in enumerate([s.semester, s.name, s.code, s.credits, s.internal_marks, s.external_marks, s.total_marks, s.grade, s.grade_points], 1):
            ws1.cell(row=row, column=col, value=v)
    for c in ws1.columns: ws1.column_dimensions[c[0].column_letter].width = 15

    # Sheet 2: Attendance
    ws2 = wb.create_sheet("Attendance")
    fill2 = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")
    for col, h in enumerate(["Subject","Total","Attended","%","Shortage"], 1):
        c = ws2.cell(row=1, column=col, value=h); c.fill = fill2; c.font = hf; c.alignment = center
    for row, a in enumerate(attendance, 2):
        subj = db.query(Subject).filter(Subject.id == a.subject_id).first()
        ws2.cell(row=row, column=1, value=subj.name if subj else f"#{a.subject_id}")
        ws2.cell(row=row, column=2, value=a.total_classes)
        ws2.cell(row=row, column=3, value=a.attended_classes)
        ws2.cell(row=row, column=4, value=f"{a.percentage}%")
        ws2.cell(row=row, column=5, value="Yes" if a.is_shortage else "No")
    for c in ws2.columns: ws2.column_dimensions[c[0].column_letter].width = 15

    # Sheet 3: Certifications
    ws3 = wb.create_sheet("Certifications")
    fill3 = PatternFill(start_color="BA7517", end_color="BA7517", fill_type="solid")
    for col, h in enumerate(["Name","Platform","Skills","URL","Date"], 1):
        c = ws3.cell(row=1, column=col, value=h); c.fill = fill3; c.font = hf; c.alignment = center
    for row, cert in enumerate(certs, 2):
        ws3.cell(row=row, column=1, value=cert.name)
        ws3.cell(row=row, column=2, value=cert.platform)
        ws3.cell(row=row, column=3, value=cert.skills_gained)
        ws3.cell(row=row, column=4, value=cert.certificate_url or "")
        ws3.cell(row=row, column=5, value=str(cert.completed_date) if cert.completed_date else "")
    for c in ws3.columns: ws3.column_dimensions[c[0].column_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={student.full_name}_report.xlsx"})

@router.get("/resume-summary")
def download_resume_summary(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    student = db.query(Student).filter(Student.user_id == current_user.id).first()
    if not student: raise HTTPException(status_code=404, detail="Profile not found")
    subjects = db.query(Subject).filter(Subject.student_id == student.id).all()
    certs = db.query(Certification).filter(Certification.student_id == student.id).all()
    cgpa = calc_cgpa(subjects)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    elements = []

    elements.append(Paragraph(student.full_name.upper(), ParagraphStyle('n', fontSize=22, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=4)))
    elements.append(Paragraph(f"{student.degree} — {student.branch}", ParagraphStyle('d', fontSize=12, alignment=TA_CENTER, textColor=colors.grey, spaceAfter=2)))
    elements.append(Paragraph(f"Roll: {student.roll_number} | CGPA: {cgpa}", ParagraphStyle('i', fontSize=11, alignment=TA_CENTER, spaceAfter=12)))
    elements.append(Spacer(1, 0.1*inch))

    # Education
    sec_s = ParagraphStyle('sec', fontSize=12, fontName='Helvetica-Bold', textColor=colors.HexColor('#534AB7'), spaceAfter=6)
    elements.append(Paragraph("EDUCATION", sec_s))
    ed = [["Semester", "Subjects", "SGPA"]]
    for sem in sorted(set(s.semester for s in subjects)):
        ss = [s for s in subjects if s.semester == sem and s.grade != "F"]
        if ss:
            tc = sum(s.credits for s in ss)
            wp = sum(s.grade_points * s.credits for s in ss)
            ed.append([f"Sem {sem}", str(len(ss)), str(round(wp/tc,2) if tc>0 else 0)])
    et = Table(ed, colWidths=[2*inch, 2.5*inch, 2*inch])
    et.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#534AB7')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTSIZE',(0,0),(-1,-1),10),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),
    ]))
    elements.append(et)
    elements.append(Spacer(1, 0.2*inch))

    # Strong subjects
    strong = [s for s in subjects if s.grade not in ["F", None] and s.total_marks >= 70]
    if strong:
        elements.append(Paragraph("STRONG SUBJECTS", sec_s))
        item_s = ParagraphStyle('item', fontSize=10, spaceAfter=3)
        for s in sorted(strong, key=lambda x: x.total_marks, reverse=True)[:5]:
            elements.append(Paragraph(f"• {s.name} — {s.grade} ({s.total_marks}/100)", item_s))
        elements.append(Spacer(1, 0.15*inch))

    # Certifications
    if certs:
        elements.append(Paragraph("CERTIFICATIONS", sec_s))
        item_s = ParagraphStyle('item2', fontSize=10, spaceAfter=3)
        for c in certs:
            elements.append(Paragraph(f"• {c.name} — {c.platform}", item_s))

    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={student.full_name}_resume.pdf"})